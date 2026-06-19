import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
from fpdf import FPDF
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def export_logs():
    """Export logs stored in Streamlit session state."""
    st.subheader("📤 Export Logs")

    if "logs" in st.session_state and st.session_state["logs"]:
        logs_df = pd.DataFrame(st.session_state["logs"], columns=["Log Entry"])
        st.dataframe(logs_df.tail(10), use_container_width=True)

        # CSV Export
        csv = logs_df.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Download Logs as CSV", csv, "system_logs.csv", "text/csv")

        # Excel Export with auto-fit
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            logs_df.to_excel(writer, index=False, sheet_name="Logs")
            ws = writer.sheets["Logs"]
            for i, col in enumerate(logs_df.columns, 1):
                max_len = max(logs_df[col].astype(str).map(len).max(), len(col))
                ws.column_dimensions[get_column_letter(i)].width = max_len + 2
        st.download_button("⬇️ Download Logs as Excel", buffer.getvalue(),
                           "system_logs.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("ℹ️ No logs available to export.")


def export_metrics():
    """Export metric history (CPU, Memory, Disk) with summary + visuals."""
    st.subheader("📤 Export Metrics")

    if "metric_history" in st.session_state and not st.session_state.metric_history.empty:
        df = st.session_state.metric_history
        st.dataframe(df.tail(10), use_container_width=True)

        # Summary statistics
        summary = {
            "Average CPU %": round(df["CPU"].mean(), 2),
            "Max CPU %": round(df["CPU"].max(), 2),
            "Average Memory %": round(df["Memory"].mean(), 2),
            "Max Memory %": round(df["Memory"].max(), 2),
            "Average Disk %": round(df["Disk"].mean(), 2),
            "Max Disk %": round(df["Disk"].max(), 2),
            "Total Records": len(df)
        }
        summary_df = pd.DataFrame(list(summary.items()), columns=["Metric", "Value"])

        # CSV Export
        combined_csv = pd.concat([summary_df, df], axis=0)
        csv = combined_csv.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Download Metrics as CSV", csv, "metrics_report.csv", "text/csv")

        # Excel Export with formatting
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
            df.to_excel(writer, index=False, sheet_name="Raw Data")
            ws = writer.sheets["Raw Data"]

            # Auto-fit columns
            for i, col in enumerate(df.columns, 1):
                max_len = max(df[col].astype(str).map(len).max(), len(col))
                ws.column_dimensions[get_column_letter(i)].width = max_len + 2

            # Conditional formatting
            for row in range(2, len(df) + 2):
                cpu = ws[f"B{row}"]
                mem = ws[f"C{row}"]
                disk = ws[f"D{row}"]
                if float(cpu.value) > 80:
                    cpu.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                if float(mem.value) > 70:
                    mem.fill = PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid")
                if float(disk.value) > 90:
                    disk.fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

        st.download_button("⬇️ Download Metrics as Excel", buffer.getvalue(),
                           "metrics_report.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Line charts
        fig, ax = plt.subplots(3, 1, figsize=(6, 8))
        df["CPU"].plot(ax=ax[0], title="CPU Usage (%)", color="blue")
        df["Memory"].plot(ax=ax[1], title="Memory Usage (%)", color="orange")
        df["Disk"].plot(ax=ax[2], title="Disk Usage (%)", color="green")
        plt.tight_layout()
        chart_buf = BytesIO()
        plt.savefig(chart_buf, format="png")
        chart_buf.seek(0)

        # Pie chart
        fig2, ax2 = plt.subplots(figsize=(5, 5))
        latest_values = [df["CPU"].iloc[-1], df["Memory"].iloc[-1], df["Disk"].iloc[-1]]
        labels = ["CPU", "Memory", "Disk"]
        colors = ["#66b3ff", "#ffcc99", "#99ff99"]
        ax2.pie(latest_values, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors)
        ax2.set_title("Resource Distribution")
        pie_buf = BytesIO()
        plt.savefig(pie_buf, format="png")
        pie_buf.seek(0)

        # Build PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 20)
        pdf.set_text_color(0, 102, 204)
        pdf.cell(200, 15, "System Metrics Report", ln=True, align="C")  # type: ignore
        pdf.set_font("Arial", "", 12)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 10, f"Generated on: {pd.Timestamp.now()}", ln=True)  # type: ignore
        pdf.ln(10)  # type: ignore

        # Summary table
        pdf.set_font("Arial", "B", 12)
        pdf.cell(0, 10, "Summary Statistics", ln=True)  # type: ignore
        pdf.set_font("Arial", "", 12)
        for k, v in summary.items():
            pdf.cell(90, 10, k, border=1)  # type: ignore
            pdf.cell(90, 10, str(v), border=1, ln=True)  # type: ignore
        pdf.ln(10)  # type: ignore

        # Charts
        pdf.cell(0, 10, "Usage Trends", ln=True)  # type: ignore
        pdf.image(chart_buf, x=10, y=None, w=180)

        # Insights
        pdf.add_page()
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "Insights", ln=True)  # type: ignore
        pdf.set_font("Arial", "", 12)

        cpu_val, mem_val, disk_val = latest_values
        if cpu_val < 50:
            pdf.multi_cell(0, 10, f"⚡ CPU usage is low at {cpu_val}%, indicating stable performance.")  # type: ignore
        elif cpu_val < 80:
            pdf.multi_cell(0, 10, f"⚡ CPU usage is moderate at {cpu_val}%, system is handling load well.")  # type: ignore
        else:
            pdf.multi_cell(0, 10, f"⚡ CPU usage is high at {cpu_val}%, potential bottleneck detected.")  # type: ignore

        if mem_val < 50:
            pdf.multi_cell(0, 10, f"💾 Memory usage is healthy at {mem_val}%.")  # type: ignore
        elif mem_val < 70:
            pdf.multi_cell(0, 10, f"💾 Memory usage is moderately high at {mem_val}%, optimization may help.")  # type: ignore
        else:
            pdf.multi_cell(0, 10, f"💾 Memory usage is critical at {mem_val}%, risk of slowdown.")  # type: ignore

        if disk_val < 70:
            pdf.multi_cell(0, 10, f"📂 Disk usage is balanced at {disk_val}%.")  # type: ignore
        else:
            pdf.multi_cell(0, 10, f"📂 Disk usage is heavy at {disk_val}%, consider cleanup or expansion.")  # type: ignore

        pdf.ln(10)  # type: ignore
        pdf.cell(0, 10, "Resource Distribution", ln=True)  # type: ignore
        pdf.image(pie_buf, x=40, y=None, w=120)

        # Finalize PDF
        pdf_bytes = pdf.output(dest="S").encode("latin-1")
